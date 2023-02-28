import openpyxl
from common.loggin_util import LoggingUtil
from config.config_file import xlsx_format

# 初始化日志工具，记录程序运行状态
logging_util_temp = LoggingUtil(__name__)
logging_util = logging_util_temp.init_logger()


# 获取测试接口集合
def get_testcases(file: str = r'testcase_data/excel/testcase_api.xlsx'):
    # 存放所有文件的测试接口数据
    testcase_data = list()

    wb = openpyxl.load_workbook(file)
    if "testcase_data" not in wb.sheetnames:
        raise TypeError("xlsx格式错误")
    else:
        sheet = wb["testcase_data"]
        for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=sheet.min_column,
                                   max_col=sheet.max_column, values_only=True):
            testcase_data_temp = dict(zip(xlsx_format, row))
            testcase_data.append(testcase_data_temp)
    # 日志系统记录测试用例
    logging_util.info(testcase_data)
    return testcase_data
